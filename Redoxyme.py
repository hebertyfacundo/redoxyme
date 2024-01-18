import tkinter as tk
from tkinter import ttk
from tkinter import *
import math
import openpyxl
from openpyxl import Workbook
import pandas as pd
import tkinter.filedialog as filedialog
from tkinter import messagebox
from sklearn.linear_model import LinearRegression
import os
import shutil
import hashlib
import matplotlib.pyplot as plt
import numpy as np
import datetime
import time
from PIL import Image, ImageTk
import clipboard
import warnings

class MainWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        self.wm_title("Redoxyme - Redox Enzyme Activity Calculator")
        self.geometry("1200x600+80+30")
        self.resizable(False,False)
        icon_path = 'Icon Art.png'
        icon = ImageTk.PhotoImage(Image.open(icon_path))
        self.iconphoto(False, icon)

        self.button1 = tk.Button(self, text="Catalase", command=self.open_catalase_window,
                                 width=15, height=2, bg="light blue", cursor="hand2", font=("TkDefaultFont", 20, "bold"))
        self.button1.grid(row = 10, column = 0, columnspan = 3, padx = 5, pady = 15)

        self.button2 = tk.Button(self, text="Glutathione Peroxidase", command=self.open_gpx_window,
                                 width=20, height=2, bg=("#F37584"), cursor="hand2", font=("TkDefaultFont", 20, "bold"))
        self.button2.grid(row = 10, column = 3, columnspan = 3, padx = 5, pady = 15)

        self.button3 = tk.Button(self, text="Superoxide Dismutase", command=self.open_sod_window,
                                 width=25, height=2, bg=("#33CC66"), cursor="hand2",
                                 font=("TkDefaultFont", 20, "bold"))
        self.button3.grid(row = 10, column = 6, columnspan = 3, padx = 5, pady = 15)

        __location__ = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__)))
        self.logo = PhotoImage(file=os.path.join(__location__, "Redox reactions.png")).subsample(5, 5)

        label = tk.Label(self, image=self.logo)
        label.image = self.logo
        label.grid(row=3, column=5, sticky="new")
        self.grid_columnconfigure(8, weight=1)
        self.grid_rowconfigure(3, weight=1)

        self.label3 = tk.Label(self, text='         REDOXYME - A Calculator for \n'
                                          '         Antioxidant Enzyme Activity \n',
                                 width=30, height=3, font=("TkDefaultFont", 20, "bold"))
        self.label3.grid(row = 1, column = 2, columnspan =6, padx = 20, pady = 5)

    def open_catalase_window(self):
        self.state("iconic")
        self.catalase_window = CatalaseWindow(self)

    def open_gpx_window(self):
        self.state("iconic")
        self.gpx_window = GpxWindow(self)

    def open_sod_window(self):
        self.state("iconic")
        self.sod_window = SodWindow(self)


class CatalaseWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Catalase Activity Calculator")
        self.geometry("1950x820")
        self.configure(background='#EFFFFE')
        icon_path = 'Icon Art.png'
        icon = ImageTk.PhotoImage(Image.open(icon_path))
        self.iconphoto(False, icon)

        frame1 = tk.LabelFrame(self, text='')
        frame1.pack(padx=1, pady=1)
        frame1.place(relx=0.5, rely=0.3, anchor='c')

        Absorbance0 = tk.StringVar()
        Absorbance60 = tk.StringVar()
        Reaction_volume = tk.StringVar()
        Sample_volume = tk.StringVar()
        mgprot = tk.StringVar()

        def answer():
            RAbs = float(Absorbance0.get()) / float(Absorbance60.get())
            LN = math.log10(RAbs) / math.log10(2.71828)
            React_volume = float(Reaction_volume.get()) / float(Sample_volume.get())
            U = LN * 0.0166666666
            U_mL = U * React_volume

            output = U_mL / float(mgprot.get())

            print(output)
            output = round(output, 4)
            Output_label.config(text=str(output))
            clipboard.copy(output)

        Absorbance0entry = tk.Entry(frame1, width=12, text = Absorbance0)
        Absorbance60entry = tk.Entry(frame1, width=12, textvariable=Absorbance60)
        Reaction_volumeentry = tk.Entry(frame1, width=12, textvariable=Reaction_volume)
        Sample_volumeentry = tk.Entry(frame1, width=12, textvariable=Sample_volume)
        mgprotentry = tk.Entry(frame1, width=12, textvariable=mgprot)

        Absorbance0_label = tk.Label(frame1, text = '  Absorbance 0 sample', font=('calibre', 20))
        Absorbance60_label = tk.Label(frame1, text='   Absorbance 60 sample', font=('calibre', 20))
        Reaction_volume_label = tk.Label(frame1, text=' Reaction Volume', font=('calibre', 20))
        Sample_volume_label = tk.Label(frame1, text='   Sample Volume', font=('calibre', 20))
        mgprotentry_label = tk.Label(frame1, text='Sample Protein (mg/mL)', font=('calibre', 20))
        Output_label = tk.Label(frame1, text='', font=('calibre', 10))
        Result_label = tk.Label(frame1, text='  Activity: U/mg Protein', font=('calibre', 15), bg=("dark gray"))

        Absorbance0_label.grid(row=2, column=1)
        Absorbance0entry.grid(row=2, column=2)
        Absorbance60_label.grid(row=3, column=1)
        Absorbance60entry.grid(row=3, column=2)
        Reaction_volume_label.grid(row=4, column=1)
        Reaction_volumeentry.grid(row=4, column=2)
        Sample_volume_label.grid(row=5, column=1)
        Sample_volumeentry.grid(row=5, column=2)
        mgprotentry_label.grid(row=6, column=1)
        mgprotentry.grid(row=6, column=2)
        Result_label.grid(row=7, column=1)
        Output_label.grid(row=7, column=2)

        calculate_Button = tk.Button(frame1, text='Calculate', command=answer, cursor="hand2")
        calculate_Button.grid(row="8", column="4")

        self.bind('<Return>', lambda event: calculate_Button.invoke())

        frame2 = tk.LabelFrame(self, text='')
        frame2.pack(padx=5, pady=5)
        frame2.place(relx=0.5, rely=0.7, anchor='c')

        list_types = ["Control", "Sample", "Treated", ""]
        Combobox_select_type1 = tk.StringVar()
        Combobox_select_type1 = ttk.Combobox(frame2, values=list_types, font=("Arial", "8", "bold"))
        Combobox_select_type1.grid(row=1, column=1, columnspan=3)
        Combobox_select_type2 = tk.StringVar()
        Combobox_select_type2 = ttk.Combobox(frame2, values=list_types, font=("Arial", "8", "bold"))
        Combobox_select_type2.grid(row=1, column=4, columnspan=3)
        Combobox_select_type3 = tk.StringVar()
        Combobox_select_type3 = ttk.Combobox(frame2, values=list_types, font=("Arial", "8", "bold"))
        Combobox_select_type3.grid(row=1, column=7, columnspan=3)
        Combobox_select_type4 = tk.StringVar()
        Combobox_select_type4 = ttk.Combobox(frame2, values=list_types, font=("Arial", "8", "bold"))
        Combobox_select_type4.grid(row=1, column=10, columnspan=3)
        Combobox_select_type5 = tk.StringVar()
        Combobox_select_type5 = ttk.Combobox(frame2, values=list_types, font=("Arial", "8", "bold"))
        Combobox_select_type5.grid(row=1, column=13, columnspan=3)
        Combobox_select_type6 = tk.StringVar()
        Combobox_select_type6 = ttk.Combobox(frame2, values=list_types, font=("Arial", "8", "bold"))
        Combobox_select_type6.grid(row=1, column=16, columnspan=3)

        frame2_sampleentry1 = tk.StringVar()
        frame2_sampleentry1 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry1.grid(row=2, column=1)
        frame2_sampleentry2 = tk.StringVar()
        frame2_sampleentry2 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry2.grid(row=2, column=4)
        frame2_sampleentry3 = tk.StringVar()
        frame2_sampleentry3 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry3.grid(row=2, column=7)
        frame2_sampleentry4 = tk.StringVar()
        frame2_sampleentry4 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry4.grid(row=2, column=10)
        frame2_sampleentry5 = tk.StringVar()
        frame2_sampleentry5 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry5.grid(row=2, column=13)
        frame2_sampleentry6 = tk.StringVar()
        frame2_sampleentry6 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry6.grid(row=2, column=16)

        frame2_sampleentry7 = tk.StringVar()
        frame2_sampleentry7 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry7.grid(row=3, column=1)
        frame2_sampleentry8 = tk.StringVar()
        frame2_sampleentry8 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry8.grid(row=3, column=4)
        frame2_sampleentry9 = tk.StringVar()
        frame2_sampleentry9 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry9.grid(row=3, column=7)
        frame2_sampleentry10 = tk.StringVar()
        frame2_sampleentry10 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry10.grid(row=3, column=10)
        frame2_sampleentry11 = tk.StringVar()
        frame2_sampleentry11 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry11.grid(row=3, column=13)
        frame2_sampleentry12 = tk.StringVar()
        frame2_sampleentry12 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry12.grid(row=3, column=16)

        frame2_sampleentry13 = tk.StringVar()
        frame2_sampleentry13 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry13.grid(row=4, column=1)
        frame2_sampleentry14 = tk.StringVar()
        frame2_sampleentry14 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry14.grid(row=4, column=4)
        frame2_sampleentry15 = tk.StringVar()
        frame2_sampleentry15 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry15.grid(row=4, column=7)
        frame2_sampleentry16 = tk.StringVar()
        frame2_sampleentry16 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry16.grid(row=4, column=10)
        frame2_sampleentry17 = tk.StringVar()
        frame2_sampleentry17 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry17.grid(row=4, column=13)
        frame2_sampleentry18 = tk.StringVar()
        frame2_sampleentry18 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry18.grid(row=4, column=16)

        frame2_sampleentry19 = tk.StringVar()
        frame2_sampleentry19 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry19.grid(row=5, column=1)
        frame2_sampleentry20 = tk.StringVar()
        frame2_sampleentry20 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry20.grid(row=5, column=4)
        frame2_sampleentry21 = tk.StringVar()
        frame2_sampleentry21 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry21.grid(row=5, column=7)
        frame2_sampleentry22 = tk.StringVar()
        frame2_sampleentry22 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry22.grid(row=5, column=10)
        frame2_sampleentry23 = tk.StringVar()
        frame2_sampleentry23 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry23.grid(row=5, column=13)
        frame2_sampleentry24 = tk.StringVar()
        frame2_sampleentry24 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry24.grid(row=5, column=16)

        frame2_sampleentry25 = tk.StringVar()
        frame2_sampleentry25 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry25.grid(row=6, column=1)
        frame2_sampleentry26 = tk.StringVar()
        frame2_sampleentry26 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry26.grid(row=6, column=4)
        frame2_sampleentry27 = tk.StringVar()
        frame2_sampleentry27 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry27.grid(row=6, column=7)
        frame2_sampleentry28 = tk.StringVar()
        frame2_sampleentry28 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry28.grid(row=6, column=10)
        frame2_sampleentry29 = tk.StringVar()
        frame2_sampleentry29 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry29.grid(row=6, column=13)
        frame2_sampleentry30 = tk.StringVar()
        frame2_sampleentry30 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry30.grid(row=6, column=16)

        frame2_sampleentry31 = tk.StringVar()
        frame2_sampleentry31 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry31.grid(row=7, column=1)
        frame2_sampleentry32 = tk.StringVar()
        frame2_sampleentry32 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry32.grid(row=7, column=4)
        frame2_sampleentry33 = tk.StringVar()
        frame2_sampleentry33 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry33.grid(row=7, column=7)
        frame2_sampleentry34 = tk.StringVar()
        frame2_sampleentry34 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry34.grid(row=7, column=10)
        frame2_sampleentry35 = tk.StringVar()
        frame2_sampleentry35 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry35.grid(row=7, column=13)
        frame2_sampleentry36 = tk.StringVar()
        frame2_sampleentry36 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry36.grid(row=7, column=16)

        widgets = [Combobox_select_type1, Combobox_select_type2, Combobox_select_type3, Combobox_select_type4,
                   Combobox_select_type5, Combobox_select_type6,
                   frame2_sampleentry1, frame2_sampleentry2, frame2_sampleentry3, frame2_sampleentry4,
                   frame2_sampleentry5, frame2_sampleentry6, frame2_sampleentry7, frame2_sampleentry8,
                   frame2_sampleentry9, frame2_sampleentry10, frame2_sampleentry11, frame2_sampleentry12,
                   frame2_sampleentry13, frame2_sampleentry14, frame2_sampleentry15, frame2_sampleentry16,
                   frame2_sampleentry17, frame2_sampleentry18, frame2_sampleentry19, frame2_sampleentry20,
                   frame2_sampleentry21, frame2_sampleentry22, frame2_sampleentry23, frame2_sampleentry24,
                   frame2_sampleentry25, frame2_sampleentry26, frame2_sampleentry27, frame2_sampleentry28,
                   frame2_sampleentry29, frame2_sampleentry30, frame2_sampleentry31, frame2_sampleentry32,
                   frame2_sampleentry33, frame2_sampleentry34, frame2_sampleentry35, frame2_sampleentry36]

        for i in range(0, len(widgets), 12):
            widget_group = widgets[i:i + 6]
            for widget in widget_group:
                widget.configure(background='light gray')


        frame3 = tk.LabelFrame(self, text='')
        frame3.pack(padx=10, pady=20)
        frame3.place(relx=0.2, rely=0.95, anchor='sw')

        def open_excel():
            file_path = filedialog.askopenfilename(initialdir="/", title="Select excel file",
                                                   filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
            if file_path.endswith(".xlsx"):
                engine = 'openpyxl'
            elif file_path.endswith(".xls"):
                engine = 'xlrd'
            else:
                raise ValueError("Invalid file extension")

            data = pd.read_excel(file_path, engine=engine)

            # Plot xy graphic using the first and second columns (A and B)
            plt.plot(data.iloc[:, 0], data.iloc[:, 1])
            plt.xlabel('Time (Secs)')
            plt.ylabel('Absorbance (A.U.)')
            plt.ion()
            plt.show()

        open_excel_button = tk.Button(frame3, text='Graph XY Excel', font=("Arial", "12"),
                                      width=16, bg="#04D4ED", fg="#000000", activebackground="#286F63",
                                      activeforeground="#D0FEF7", command=open_excel, cursor="hand2")
        open_excel_button.grid(row=1, column=1)

        def plot():
            warnings.filterwarnings("ignore", category=RuntimeWarning)
            group1 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry1, frame2_sampleentry7, frame2_sampleentry13, frame2_sampleentry19,
                       frame2_sampleentry25, frame2_sampleentry31]]

            group2 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry2, frame2_sampleentry8, frame2_sampleentry14, frame2_sampleentry20,
                       frame2_sampleentry26, frame2_sampleentry32]]

            group3 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry3, frame2_sampleentry9, frame2_sampleentry15, frame2_sampleentry21,
                       frame2_sampleentry27, frame2_sampleentry33]]

            group4 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry4, frame2_sampleentry10, frame2_sampleentry16, frame2_sampleentry22,
                       frame2_sampleentry28, frame2_sampleentry34]]

            group5 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry5, frame2_sampleentry11, frame2_sampleentry17, frame2_sampleentry23,
                       frame2_sampleentry29, frame2_sampleentry35]]

            group6 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry6, frame2_sampleentry12, frame2_sampleentry18, frame2_sampleentry24,
                       frame2_sampleentry30, frame2_sampleentry36]]

            def stddev(data):
                n = len(data)
                if n == 0:
                    return np.nan  # Return NaN for standard deviation if the group is empty
                mean_value = sum(data) / n
                variance = sum((x - mean_value) ** 2 for x in data) / (n - 1)
                return np.sqrt(variance)

            group1_filtered = [x for x in group1 if x is not None and not np.isnan(x)]
            group1_mean = np.mean(group1_filtered)
            group1_sd = stddev(group1_filtered)

            group2_filtered = [x for x in group2 if x is not None and not np.isnan(x)]
            group2_mean = np.mean(group2_filtered)
            group2_sd = stddev(group2_filtered)

            group3_filtered = [x for x in group3 if x is not None and not np.isnan(x)]
            group3_mean = np.mean(group3_filtered)
            group3_sd = stddev(group3_filtered)

            group4_filtered = [x for x in group4 if x is not None and not np.isnan(x)]
            group4_mean = np.mean(group4_filtered)
            group4_sd = stddev(group4_filtered)

            group5_filtered = [x for x in group5 if x is not None and not np.isnan(x)]
            group5_mean = np.mean(group5_filtered)
            group5_sd = stddev(group5_filtered)

            group6_filtered = [x for x in group6 if x is not None and not np.isnan(x)]
            group6_mean = np.mean(group6_filtered)
            group6_sd = stddev(group6_filtered)

            fig, ax = plt.subplots()

            bar1 = ax.bar(1, group1_mean, yerr=group1_sd, label=Combobox_select_type1.get(), ecolor='red', capsize=5)
            bar2 = ax.bar(2, group2_mean, yerr=group2_sd, label=Combobox_select_type2.get(), ecolor='red', capsize=5)
            bar3 = ax.bar(3, group3_mean, yerr=group3_sd, label=Combobox_select_type3.get(), ecolor='red', capsize=5)
            bar4 = ax.bar(4, group4_mean, yerr=group4_sd, label=Combobox_select_type4.get(), ecolor='red', capsize=5)
            bar5 = ax.bar(5, group5_mean, yerr=group5_sd, label=Combobox_select_type5.get(), ecolor='red', capsize=5)
            bar6 = ax.bar(6, group6_mean, yerr=group6_sd, label=Combobox_select_type6.get(), ecolor='red', capsize=5)

            ax.set_ylabel('Catalase Activity (U/mg protein)')
            ax.set_xticks([1, 2, 3, 4, 5, 6])
            ax.set_xticklabels(
                [(Combobox_select_type1.get()), (Combobox_select_type2.get()), (Combobox_select_type3.get()),
                 (Combobox_select_type4.get()), (Combobox_select_type5.get()), (Combobox_select_type6.get())])

            plt.show()

        plot_button = tk.Button(frame3, text='Plot', font=("Arial", "12"),
                                width=16, bg="#04D4ED", fg="#000000", activebackground="#286F63",
                                activeforeground="#D0FEF7", command=plot, cursor="hand2")
        plot_button.grid(row=1, column=2)
        warnings.resetwarnings()

        def clear():

                result = messagebox.askyesno("Save Changes",
                                             "ARE YOU SURE YOU WANT TO CLEAR? That will erase all unsaved data! ")
                if result == True:
                    for widget in widgets:
                        if isinstance(widget, tk.Entry):
                            widget.delete(0, 300)
                        elif isinstance(widget, tk.ttk.Combobox):
                            widget.set("")

        clear_button = tk.Button(frame3, text='Clear', font=("Arial", "12"),
                                 width=16, bg="#04D4ED", fg="#000000", activebackground="#286F63",
                                 activeforeground="#D0FEF7", command=clear)
        clear_button.grid(row=1, column=3)

        def save_member():
            file_name = "Catalaseactivity_" + datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + ".xlsx"
            file = openpyxl.Workbook()
            sheet = file.active

            rows = 7
            cols = 6
            for row in range(rows):
                for col in range(cols):
                    i = row * cols + col
                    sheet.cell(column=col + 2, row=row + 3, value=widgets[i].get())

            file.save(file_name)
            file.close()

        save_button = tk.Button(frame3, text='Save', font=("Arial", "12"),
                                width=15, bg="#04D4ED", fg="#000000", activebackground="#286F63",
                                activeforeground="#D0FEF7", command=save_member, cursor="hand2")
        save_button.grid(row= 1, column = 4)

        def openFile():
            the_file = filedialog.askopenfilename(title="Select the updated file", filetypes=[("All files", "*.*")])
            if the_file:
                print(f"Selected file: {the_file}")
                try:
                    os.startfile(os.path.abspath(the_file))
                except Exception as e:
                    print(f"Error opening file: {e}")
            else:
                self.state("iconic")
        open_button = tk.Button(frame3, text="Open File", font=("Arial", "12"),
                                width=18, bg="#04D4ED", fg="#000000", activebackground="#286F63",
                                activeforeground="#D0FEF7", command=openFile, cursor="hand2")
        open_button.grid(row= 1, column = 5)


        def open_window():
            window2 = tk.Toplevel()
            window2.title('Instructions')
            window2.geometry('6000x3000')
            icon_path = 'Icon Art.png'
            icon = ImageTk.PhotoImage(Image.open(icon_path))
            window2.iconphoto(False, icon)

            label_name = tk.Label(window2,
                                  text='Absorbance 0 sample → initial absorbance of H2O2, which should be close to 0.5. \n'
                                       'Absorbance 60 sample → absorbance at 60 seconds. \n'
                                       'Reaction Volume → total volume in the cuvette. \n'
                                       'Samples Volume → volume of samples in the cuvette. \n'
                                       'mg protein/mL → protein concentration of the original sample.\n'
                                       '\n'
                                       'Formula:\n'
                                       'U = (1/60) x ln (Absorbance 0 sample/Absorbance 60 sample)\n'
                                       'U total/mL = U x (Reaction volume/Sample volume) \n'
                                       'U/mg protein = (U total/mL)/(concentration of sample protein in mg/mL) \n'
                                       'Results are in U/mg protein. All experiments are conducted at 240 nm.\n'
                                       'The result is automatically copied to clipboard so the user may paste in the table provided or in another software\n'
                                       'One unit will decompose 1.0 μmol of H2O2 per min at pH 7.0 at 25 °C\n'
                                       'Users are expected to enter the data in the table and choose or enter the names or codes of the samples.\n'
                                       'Each cell in the table can accept numbers, letters, or symbols. The plot function will not work with letters or symbols\n'
                                       '\n'
                                       'When pressed, the SAVE BUTTON will save a file in the same folder as the REDOXYME.exe file. \n'
                                       'The file will be saved in Excel format (named "catalaseactivity year-month-day hour-minute-second.xlsx"). The user should save it in a different folder.\n'
                                       'If the SAVE BUTTON is pressed again, it will SAVE IT AGAIN with the current information of Y-M-D H-min-sec.\n'
                                       'The Plot button will create a plot from data. The Graph XY excel button is used to open an excel with traces of enzyme activity.\n'
                                       '\n'
                                       'The user can open the above-saved file or any other file on their computer by pressing the OPEN FILE button. \n'
                                       '\n'
                                       'The CLEAR button, when pressed, will erase all data. Please, check for proper data in the saved file. \n'
                                       '\nContact email: heberty.facundo@gmail.com',
                                  bd=1, justify='left', font=('calibre', 16))

            label_name.grid(row=4, column=0, sticky='w')
            button_voltar = tk.Button(window2, text='Close', font=('calibre', 20), command=window2.destroy, cursor="hand2")
            button_voltar.place(x=1050, y=600)

        Window2button1 = tk.Button(self, text='Intructions', command=open_window, cursor="hand2")
        Window2button1.grid(row=2, column=3)

        def window_prot1():
            window_prot1 = tk.Toplevel()
            window_prot1.title('Protein')
            window_prot1.geometry('1200x300')
            icon_path = 'Icon Art.png'
            icon = ImageTk.PhotoImage(Image.open(icon_path))
            window_prot1.iconphoto(False, icon)


            reg = LinearRegression()

            x_label = tk.Label(window_prot1, text="Protein_Concentration (mg/mL)(X)")
            x_label.grid(row=0, column=0, sticky="W")

            x_entries = [tk.Entry(window_prot1) for i in range(7)]
            for i, entry in enumerate(x_entries):
                entry.grid(row=0, column=i + 1)

            y_label = tk.Label(window_prot1, text="Absorbance (A.U) (Y)")
            y_label.grid(row=1, column=0, sticky="W")

            y_entries = [tk.Entry(window_prot1) for i in range(7)]
            for i, entry in enumerate(y_entries):
                entry.grid(row=1, column=i + 1)

            slo_entry = tk.Entry(window_prot1)
            slo_entry.grid(row=3, column=2, sticky="W")

            int_entry = tk.Entry(window_prot1)
            int_entry.grid(row=4, column=2, sticky="W")

            def calculate():
                x_values = [entry.get() for entry in x_entries]
                y_values = [entry.get() for entry in y_entries]

                x = []
                y = []
                for i in range(len(x_values)):
                    if x_values[i] and y_values[i]:
                        x.append(float(x_values[i]))
                        y.append(float(y_values[i]))

                if x and y:
                    x = np.array(x).reshape(-1, 1)
                    y = np.array(y)

                    reg.fit(x, y)

                    slope = reg.coef_[0]
                    intercept = reg.intercept_
                    r_squared = reg.score(x, y)

                    slope_value_label.config(text=str(slope))
                    intercept_value_label.config(text=str(intercept))

                    slo_entry.delete(0, 'end')
                    slo_entry.insert(0, str(slope))

                    int_entry.delete(0, 'end')
                    int_entry.insert(0, str(intercept))

                    plt.scatter(x, y)
                    plt.plot(x, reg.predict(x), color='Red')
                    plt.xlabel("Protein Concentration (mg/mL)")
                    plt.ylabel("Absorbance")
                    plt.text(0.05, 0.95, 'R² = {:.3f}'.format(r_squared),
                             transform=plt.gca().transAxes, fontsize=12,
                             verticalalignment='top')
                    plt.show()

                else:
                    messagebox.showerror("Error", "Did you forget the numbers? Decimals must be separated by dot! Do not use comma.")

            calculate_button = tk.Button(window_prot1, text="Calculate", command=calculate)
            calculate_button.grid(row=2, column=1, pady=10)

            slope_label = tk.Label(window_prot1, text="Slope:")
            slope_label.grid(row=3, column=0, sticky="W")

            slope_value_label = tk.Label(window_prot1)
            slope_value_label.grid(row=3, column=1)

            intercept_label = tk.Label(window_prot1, text="Intercept:")
            intercept_label.grid(row=4, column=0, sticky="W")


            intercept_value_label = tk.Label(window_prot1)
            intercept_value_label.grid(row=4, column=1)

            slope_label = tk.Label(window_prot1, text="Sample Absorbance")
            slope_label.grid(row=3, column=4, sticky="W")

            sample_entry = tk.Entry(window_prot1)
            sample_entry.grid(row=4, column=4, sticky="W")

            def sample():
                first = float(sample_entry.get()) - float(int_entry.get())
                second = float(slo_entry.get())
                output = first / second
                print(output)
                output = round(output, 4)
                output_label.config(text=str(output))
                clipboard.copy(output)

            output_label = tk.Label(window_prot1, text='', font=('calibre', 10))
            output_label.grid(row=8, column=5, sticky="W")
            Result_label = tk.Label(window_prot1, text='mg Protein/mL', font=('calibre', 15), bg=("dark gray"))
            Result_label.grid(row=8, column=4, sticky="W")

            Sample_Button = tk.Button(window_prot1, text='Calculate Protein', command=sample)
            Sample_Button.grid(row="4", column="5")

            button_voltar = tk.Button(window_prot1, text='Close', font=('calibre', 20), command=window_prot1.destroy, cursor="hand2")
            button_voltar.place(x=1050, y=600)

        Window_prot1_button1 = tk.Button(self, text='Protein', command=window_prot1, cursor="hand2")
        Window_prot1_button1.grid(row=3, column=3)

class GpxWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Glutathione Peroxidase Activity Calculator")
        self.geometry("2000x820")
        self.configure(background='#FCEDEA')
        icon_path = 'Icon Art.png'
        icon = ImageTk.PhotoImage(Image.open(icon_path))
        self.iconphoto(False, icon)

        frame1 = tk.LabelFrame(self, text='')
        frame1.pack(padx=1, pady=1)
        frame1.place(relx=0.5, rely=0.3, anchor='c')

        Abs0sample =tk.StringVar()
        Abs0blank=tk.StringVar()
        Abs60sample=tk.StringVar()
        Abs60blank=tk.StringVar()
        Reaction_volume=tk.StringVar()
        Sample_volume=tk.StringVar()
        Dilutionfactor=tk.StringVar()
        Coef_Ext = tk.StringVar()
        mgprot=tk.StringVar()


        def answer():
            DeltAbs_sample = float(Abs0sample.get()) - float(Abs60sample.get())
            DeltAbs_blank = float(Abs0blank.get()) - float(Abs60blank.get())
            Delta_delta = (DeltAbs_sample-DeltAbs_blank)
            GPX_1 = Delta_delta/float(Coef_Ext.get())
            Coef_1 = float(Reaction_volume.get()) / float(Sample_volume.get())
            Coef_2 = Coef_1*float(Dilutionfactor.get())*2
            GPX_2 = GPX_1*Coef_2
            mgprotsample = float(mgprot.get())
            output = GPX_2/mgprotsample
            output = round(output, 4)
            Output_label.config(text=str(output))

            clipboard.copy(output)


        Abs0sampleentry = tk.Entry(frame1, width=12, textvariable = Abs0sample)
        Abs60sampleentry = tk.Entry(frame1, width=12, textvariable = Abs60sample)
        Abs0blankentry = tk.Entry(frame1, width=12, textvariable = Abs0blank)
        Abs60blankentry = tk.Entry(frame1, width=12, textvariable = Abs60blank)
        Reaction_volumeentry = tk.Entry(frame1, width=12, textvariable = Reaction_volume)
        Sample_volumeentry = tk.Entry(frame1, width=12, textvariable = Sample_volume)
        Dilutionfactorentry= tk.Entry(frame1, width=12, textvariable = Dilutionfactor)
        Coef_Extentry = tk.Entry(frame1, width=12, textvariable = Coef_Ext)
        mgprotentry= tk.Entry(frame1, width=12, textvariable = mgprot)


        Abs0sample_label=tk.Label(frame1, text = '       Absorbance 0 sample', font=('calibre',20))
        Abs60sample_label=tk.Label(frame1, text = '     Absorbance 60 sample', font=('calibre',20))
        Abs0blank_label=tk.Label(frame1, text = '         Absorbance 0 blank', font=('calibre',20))
        Abs60blank_label=tk.Label(frame1, text = '       Absorbance 60 blank', font=('calibre',20))
        Reaction_volume_label = tk.Label(frame1, text = '             Reaction Volume', font=('calibre',20))
        Sample_volume_label = tk.Label(frame1, text = '               Sample Volume', font=('calibre',20))
        Dilutionfactor_label= tk.Label(frame1, text = '                Dilution Factor', font=('calibre',20))
        Coef_Ext_label = tk.Label(frame1, text = '        Coeficient Extinction', font=('calibre',20))
        mgprot_label= tk.Label(frame1, text = '                 mg protein/mL', font=('calibre',20))
        Output_label = tk.Label(frame1, text = '', font=('calibre',10))
        Result_label = tk.Label(frame1, text = '                  Activity: U/mg Protein', font=('calibre',15), bg=("dark gray"))


        Abs0sample_label.grid(row=2, column=1)
        Abs0sampleentry.grid(row=2, column=2)
        Abs60sample_label.grid(row=3, column=1)
        Abs60sampleentry.grid(row=3, column=2)
        Abs0blank_label.grid(row=4, column=1)
        Abs0blankentry.grid(row=4, column=2)
        Abs60blank_label.grid(row=5, column=1)
        Abs60blankentry.grid(row=5, column=2)
        Reaction_volume_label.grid(row=6, column=1)
        Reaction_volumeentry.grid(row=6, column=2)
        Sample_volume_label.grid(row=7, column=1)
        Sample_volumeentry.grid(row=7, column=2)
        Dilutionfactor_label.grid(row=8, column=1)
        Dilutionfactorentry.grid(row=8, column=2)
        Coef_Ext_label.grid(row=9, column=1)
        Coef_Extentry.grid(row=9, column=2)
        mgprot_label.grid(row=10, column=1)
        mgprotentry.grid(row=10, column=2)
        Result_label.grid(row=11, column=1)
        Output_label.grid(row=11, column=2)

        calculate_Button = tk.Button(frame1, text='Calculate', command=answer, cursor="hand2")
        calculate_Button.grid(row="12", column="4")

        self.bind('<Return>', lambda event: calculate_Button.invoke())

        frame2 = tk.LabelFrame(self, text='')
        frame2.pack(padx=5, pady=5)
        frame2.place(relx=0.5, rely=0.75, anchor='c')

        list_types = ["Control", "Sample", "Treated", ""]
        Combobox_select_type1 = tk.StringVar()
        Combobox_select_type1 = ttk.Combobox(frame2,values=list_types, font = ("Arial", "8", "bold"))
        Combobox_select_type1.grid(row=1, column=1, columnspan= 3)
        Combobox_select_type2 = tk.StringVar()
        Combobox_select_type2 = ttk.Combobox(frame2,values=list_types, font = ("Arial", "8", "bold"))
        Combobox_select_type2.grid(row=1, column=4, columnspan= 3)
        Combobox_select_type3 = tk.StringVar()
        Combobox_select_type3 = ttk.Combobox(frame2,values=list_types, font = ("Arial", "8", "bold"))
        Combobox_select_type3.grid(row=1, column=7, columnspan= 3)
        Combobox_select_type4 = tk.StringVar()
        Combobox_select_type4 = ttk.Combobox(frame2,values=list_types, font = ("Arial", "8", "bold"))
        Combobox_select_type4.grid(row=1, column=10, columnspan= 3)
        Combobox_select_type5 = tk.StringVar()
        Combobox_select_type5 = ttk.Combobox(frame2,values=list_types, font = ("Arial", "8", "bold"))
        Combobox_select_type5.grid(row=1, column=13, columnspan= 3)
        Combobox_select_type6 = tk.StringVar()
        Combobox_select_type6 = ttk.Combobox(frame2,values=list_types, font = ("Arial", "8", "bold"))
        Combobox_select_type6.grid(row=1, column=16, columnspan= 3)

        frame2_sampleentry1 = tk.StringVar()
        frame2_sampleentry1 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry1.grid(row = 2, column = 1)
        frame2_sampleentry2 = tk.StringVar()
        frame2_sampleentry2 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry2.grid(row = 2, column = 4)
        frame2_sampleentry3 = tk.StringVar()
        frame2_sampleentry3 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry3.grid(row = 2, column = 7)
        frame2_sampleentry4 = tk.StringVar()
        frame2_sampleentry4 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry4.grid(row = 2, column = 10)
        frame2_sampleentry5 = tk.StringVar()
        frame2_sampleentry5 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry5.grid(row = 2, column = 13)
        frame2_sampleentry6 = tk.StringVar()
        frame2_sampleentry6 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry6.grid(row = 2, column = 16)

        frame2_sampleentry7 = tk.StringVar()
        frame2_sampleentry7 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry7.grid(row = 3, column = 1)
        frame2_sampleentry8 = tk.StringVar()
        frame2_sampleentry8 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry8.grid(row = 3, column = 4)
        frame2_sampleentry9 = tk.StringVar()
        frame2_sampleentry9 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry9.grid(row = 3, column = 7)
        frame2_sampleentry10 = tk.StringVar()
        frame2_sampleentry10 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry10.grid(row = 3, column = 10)
        frame2_sampleentry11 = tk.StringVar()
        frame2_sampleentry11 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry11.grid(row = 3, column = 13)
        frame2_sampleentry12 = tk.StringVar()
        frame2_sampleentry12 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry12.grid(row = 3, column = 16)

        frame2_sampleentry13 = tk.StringVar()
        frame2_sampleentry13 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry13.grid(row = 4, column = 1)
        frame2_sampleentry14 = tk.StringVar()
        frame2_sampleentry14 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry14.grid(row = 4, column = 4)
        frame2_sampleentry15 = tk.StringVar()
        frame2_sampleentry15 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry15.grid(row = 4, column = 7)
        frame2_sampleentry16 = tk.StringVar()
        frame2_sampleentry16 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry16.grid(row = 4, column = 10)
        frame2_sampleentry17 = tk.StringVar()
        frame2_sampleentry17 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry17.grid(row = 4, column = 13)
        frame2_sampleentry18 = tk.StringVar()
        frame2_sampleentry18 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry18.grid(row = 4, column = 16)

        frame2_sampleentry19 = tk.StringVar()
        frame2_sampleentry19 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry19.grid(row = 5, column = 1)
        frame2_sampleentry20 = tk.StringVar()
        frame2_sampleentry20 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry20.grid(row = 5, column = 4)
        frame2_sampleentry21 = tk.StringVar()
        frame2_sampleentry21 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry21.grid(row = 5, column = 7)
        frame2_sampleentry22 = tk.StringVar()
        frame2_sampleentry22 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry22.grid(row = 5, column = 10)
        frame2_sampleentry23 = tk.StringVar()
        frame2_sampleentry23 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry23.grid(row = 5, column = 13)
        frame2_sampleentry24 = tk.StringVar()
        frame2_sampleentry24 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry24.grid(row = 5, column = 16)

        frame2_sampleentry25 = tk.StringVar()
        frame2_sampleentry25 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry25.grid(row = 6, column = 1)
        frame2_sampleentry26 = tk.StringVar()
        frame2_sampleentry26 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry26.grid(row = 6, column = 4)
        frame2_sampleentry27 = tk.StringVar()
        frame2_sampleentry27 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry27.grid(row = 6, column = 7)
        frame2_sampleentry28 = tk.StringVar()
        frame2_sampleentry28 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry28.grid(row = 6, column = 10)
        frame2_sampleentry29 = tk.StringVar()
        frame2_sampleentry29 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry29.grid(row = 6, column = 13)
        frame2_sampleentry30 = tk.StringVar()
        frame2_sampleentry30 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry30.grid(row = 6, column = 16)
        frame2_sampleentry31 = tk.StringVar()
        frame2_sampleentry31 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry31.grid(row = 7, column = 1)
        frame2_sampleentry32 = tk.StringVar()
        frame2_sampleentry32 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry32.grid(row = 7, column = 4)
        frame2_sampleentry33 = tk.StringVar()
        frame2_sampleentry33 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry33.grid(row = 7, column = 7)
        frame2_sampleentry34 = tk.StringVar()
        frame2_sampleentry34 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry34.grid(row = 7, column = 10)
        frame2_sampleentry35 = tk.StringVar()
        frame2_sampleentry35 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry35.grid(row = 7, column = 13)
        frame2_sampleentry36 = tk.StringVar()
        frame2_sampleentry36 = tk.Entry(frame2, width=20,font = (15))
        frame2_sampleentry36.grid(row = 7, column = 16)

        frame3 = tk.LabelFrame(self, text='')
        frame3.pack(padx=10, pady=20)
        frame3.place(relx=0.2, rely=0.95, anchor='sw')

        def open_excel():
            file_path = filedialog.askopenfilename(initialdir="/", title="Select excel file",
                                                   filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
            if file_path.endswith(".xlsx"):
                engine = 'openpyxl'
            elif file_path.endswith(".xls"):
                engine = 'xlrd'
            else:
                raise ValueError("Invalid file extension")

            data = pd.read_excel(file_path, engine=engine)

            # Plot xy graphic using the first and second columns (A and B)
            plt.plot(data.iloc[:, 0], data.iloc[:, 1])
            plt.xlabel('Time (Secs)')
            plt.ylabel('Absorbance (A.U.)')
            plt.ion()
            plt.show()

        open_excel_button = tk.Button(frame3, text='Graph XY Excel', font=("Arial", "12"),
                                      width=16, bg="#d17486", fg="#000000", activebackground="#286F63",
                                      activeforeground="#D0FEF7", command=open_excel, cursor="hand2")
        open_excel_button.grid(row=1, column=1)

        def plot():
            warnings.filterwarnings("ignore", category=RuntimeWarning)
            group1 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry1, frame2_sampleentry7, frame2_sampleentry13, frame2_sampleentry19,
                       frame2_sampleentry25, frame2_sampleentry31]]

            group2 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry2, frame2_sampleentry8, frame2_sampleentry14, frame2_sampleentry20,
                       frame2_sampleentry26, frame2_sampleentry32]]

            group3 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry3, frame2_sampleentry9, frame2_sampleentry15, frame2_sampleentry21,
                       frame2_sampleentry27, frame2_sampleentry33]]

            group4 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry4, frame2_sampleentry10, frame2_sampleentry16, frame2_sampleentry22,
                       frame2_sampleentry28, frame2_sampleentry34]]

            group5 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry5, frame2_sampleentry11, frame2_sampleentry17, frame2_sampleentry23,
                       frame2_sampleentry29, frame2_sampleentry35]]

            group6 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry6, frame2_sampleentry12, frame2_sampleentry18, frame2_sampleentry24,
                       frame2_sampleentry30, frame2_sampleentry36]]

            def stddev(data):
                n = len(data)
                if n == 0:
                    return np.nan  # Return NaN for standard deviation if the group is empty
                mean_value = sum(data) / n
                variance = sum((x - mean_value) ** 2 for x in data) / (n - 1)
                return np.sqrt(variance)

            group1_filtered = [x for x in group1 if x is not None and not np.isnan(x)]
            group1_mean = np.mean(group1_filtered)
            group1_sd = stddev(group1_filtered)

            group2_filtered = [x for x in group2 if x is not None and not np.isnan(x)]
            group2_mean = np.mean(group2_filtered)
            group2_sd = stddev(group2_filtered)

            group3_filtered = [x for x in group3 if x is not None and not np.isnan(x)]
            group3_mean = np.mean(group3_filtered)
            group3_sd = stddev(group3_filtered)

            group4_filtered = [x for x in group4 if x is not None and not np.isnan(x)]
            group4_mean = np.mean(group4_filtered)
            group4_sd = stddev(group4_filtered)

            group5_filtered = [x for x in group5 if x is not None and not np.isnan(x)]
            group5_mean = np.mean(group5_filtered)
            group5_sd = stddev(group5_filtered)

            group6_filtered = [x for x in group6 if x is not None and not np.isnan(x)]
            group6_mean = np.mean(group6_filtered)
            group6_sd = stddev(group6_filtered)

            fig, ax = plt.subplots()

            bar1 = ax.bar(1, group1_mean, yerr=group1_sd, label=Combobox_select_type1.get(), ecolor='red', capsize=5)
            bar2 = ax.bar(2, group2_mean, yerr=group2_sd, label=Combobox_select_type2.get(), ecolor='red', capsize=5)
            bar3 = ax.bar(3, group3_mean, yerr=group3_sd, label=Combobox_select_type3.get(), ecolor='red', capsize=5)
            bar4 = ax.bar(4, group4_mean, yerr=group4_sd, label=Combobox_select_type4.get(), ecolor='red', capsize=5)
            bar5 = ax.bar(5, group5_mean, yerr=group5_sd, label=Combobox_select_type5.get(), ecolor='red', capsize=5)
            bar6 = ax.bar(6, group6_mean, yerr=group6_sd, label=Combobox_select_type6.get(), ecolor='red', capsize=5)

            ax.set_ylabel('Catalase Activity (U/mg protein)')
            ax.set_xticks([1, 2, 3, 4, 5, 6])
            ax.set_xticklabels(
                [(Combobox_select_type1.get()), (Combobox_select_type2.get()), (Combobox_select_type3.get()),
                 (Combobox_select_type4.get()), (Combobox_select_type5.get()), (Combobox_select_type6.get())])

            plt.show()

        plot_button = tk.Button(frame3, text='Plot', font=("Arial", "12"),
                                width=16, bg="#d17486", fg="#000000", activebackground="#286F63",
                                activeforeground="#D0FEF7", command=plot, cursor="hand2")
        plot_button.grid(row=1, column=2)
        warnings.resetwarnings()


        def clear():
            clear_combobox = [Combobox_select_type1, Combobox_select_type2, Combobox_select_type3,
                              Combobox_select_type4, Combobox_select_type5, Combobox_select_type6,
                              frame2_sampleentry1, frame2_sampleentry2, frame2_sampleentry3, frame2_sampleentry4,
                              frame2_sampleentry5, frame2_sampleentry6, frame2_sampleentry7, frame2_sampleentry8,
                              frame2_sampleentry9, frame2_sampleentry10, frame2_sampleentry11, frame2_sampleentry12,
                              frame2_sampleentry13, frame2_sampleentry14, frame2_sampleentry15, frame2_sampleentry16,
                              frame2_sampleentry17, frame2_sampleentry18, frame2_sampleentry19, frame2_sampleentry20,
                              frame2_sampleentry21, frame2_sampleentry22, frame2_sampleentry23, frame2_sampleentry24,
                              frame2_sampleentry25, frame2_sampleentry26, frame2_sampleentry27, frame2_sampleentry28,
                              frame2_sampleentry29, frame2_sampleentry30, frame2_sampleentry31, frame2_sampleentry32,
                              frame2_sampleentry33, frame2_sampleentry34, frame2_sampleentry35, frame2_sampleentry36]

            result = messagebox.askyesno("Save Changes",
                                         "ARE YOU SURE YOU WANT TO CLEAR? That will erase all unsaved data! ")
            if result == True:
                for widget in clear_combobox:
                    if isinstance(widget, tk.Entry):
                        widget.delete(0, 300)
                    elif isinstance(widget, tk.ttk.Combobox):
                        widget.set("")


        clear_button = tk.Button(frame3, text='Clear', font=("Arial", "12"),
                                 width=18, bg="#d17486", fg="#000000", activebackground="#286F63",
                                 activeforeground="#D0FEF7", command=clear)
        clear_button.grid(row=1, column=3)


        def save_member():
            widgets = [Combobox_select_type1, Combobox_select_type2, Combobox_select_type3, Combobox_select_type4,
                       Combobox_select_type5, Combobox_select_type6,
                       frame2_sampleentry1, frame2_sampleentry2, frame2_sampleentry3, frame2_sampleentry4, frame2_sampleentry5,
                       frame2_sampleentry6, frame2_sampleentry7, frame2_sampleentry8, frame2_sampleentry9, frame2_sampleentry10,
                       frame2_sampleentry11, frame2_sampleentry12, frame2_sampleentry13, frame2_sampleentry14,
                       frame2_sampleentry15, frame2_sampleentry16, frame2_sampleentry17, frame2_sampleentry18,
                       frame2_sampleentry19, frame2_sampleentry20, frame2_sampleentry21, frame2_sampleentry22,
                       frame2_sampleentry23, frame2_sampleentry24, frame2_sampleentry25, frame2_sampleentry26,
                       frame2_sampleentry27, frame2_sampleentry28, frame2_sampleentry29, frame2_sampleentry30,
                       frame2_sampleentry31, frame2_sampleentry32, frame2_sampleentry33, frame2_sampleentry34,
                       frame2_sampleentry35, frame2_sampleentry36]

            file_name = "GPXactivity_" + datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + ".xlsx"
            file = openpyxl.Workbook()
            sheet = file.active

            rows = 7
            cols = 6
            for row in range(rows):
                for col in range(cols):
                    i = row * cols + col
                    sheet.cell(column=col + 2, row=row + 3, value=widgets[i].get())

            file.save(file_name)
            file.close()


        save_button = tk.Button(frame3, text='Save', font=("Arial", "12"),
                                width=18, bg="#d17486", fg="#000000", activebackground="#286F63",
                                activeforeground="#D0FEF7", command=save_member, cursor="hand2")
        save_button.grid(row=1, column=4)


        def openFile():
            the_file = filedialog.askopenfilename(title="Select the updated file", filetypes=[("All files", "*.*")])
            if the_file:
                print(f"Selected file: {the_file}")
                try:
                    os.startfile(os.path.abspath(the_file))
                except Exception as e:
                    print(f"Error opening file: {e}")
            else:
                self.state("iconic")


        open_button = tk.Button(frame3, text="Open File", font=("Arial", "12"),
                                width=18, bg="#d17486", fg="#000000", activebackground="#286F63",
                                activeforeground="#D0FEF7", command=openFile, cursor="hand2")
        open_button.grid(row=1, column=5)
        def open_window():
            window2 = tk.Toplevel()
            window2.title('Instructions')
            window2.geometry('6000x3000')
            icon_path = 'Icon Art.png'
            icon = ImageTk.PhotoImage(Image.open(icon_path))
            window2.iconphoto(False, icon)

            label_name=tk.Label(window2, text='1 unit of GPx-1 is defined as the amount of enzyme necessary to catalyze the oxidation  \n'
                                      'of 1.0 micromole GSH (by H2O2) to GSSG, per minute at 25 ºC, pH 7.0.\n'
                                      'Absorbance 0 Sample → initial absorbance of NAD(P)H \n'
                                      'Absorbance 60 Sample → final absorbance \n'
                                      'Absorbance 0 blank→ initial absorbance of NAD(P)H \n'
                                      'Absorbance 60 blank→ final absorbance of NAD(P)H \n'
                                      'Reaction Volume → total volume in the cuvette \n'                                        
                                      'Samples Volume → volume of samples in the cuvette \n'
                                      'Dilution Factor → dilution of sample before adding to the cuvette \n'
                                      'Coefficient extinction of NADPH at 340 nm at 1 cm pathlength is 6.22. If the pathlength is 0.5 consider 3.11 \n'
                                      'mg protein/mL → protein concentration of the original sample\n'
                                      '\n'
                                      'U/mL = U x (Reaction volume/Sample volume) \n'
                                      'U/mg protein = (U total/mL)/(concentration of sample protein in mg/mL) \n'
                                      'Results are in U/mg protein or mU/mg protein. All experiments are conducted at 340 nm.\n'
                                              '\n' 
                                      'The result is automatically copied to clipboard so the user may paste in the table provided or in another software.\n'
                                      'SAVING DATA:\n' 
                                      'Users are expected to enter the data in the table and choose or enter the names or codes of the samples.\n'
                                      'Each cell in the table can accept numbers, letters, or symbols. The plot function will not work with letters or symbols.\n'
                                      '\n'
                                      'When pressed, the SAVE BUTTON will save a file in the same folder as the REDOXYME file. \n'
                                      'The file will be saved in Excel format (named "GPXactivity year-month-day hour-minute-second.xlsx"). \n'
                                      'If the SAVE BUTTON is pressed again, it will SAVE IT AGAIN with the current information of Y-M-D H-min-sec.\n'
                                      'The Plot button will create a plot from data. The Graph XY excel button is used to open an excel with traces of enzyme activity.\n'
                                      'The user can open the above-saved file or any other file on their computer by pressing the OPEN FILE button. \n' 
                                      '\n'
                                      'The CLEAR button, when pressed, will erase all data. Please, check for proper data in the saved file. \n'
                                      '\nContact email: heberty.facundo@gmail.com',
                             bd=1, justify='left', font=('calibre',14))

            label_name.grid(row=4, column=0, sticky='w')
            button_voltar = tk.Button(window2, text = 'Close', font=('calibre',20), command = window2.destroy, cursor="hand2")
            button_voltar.place(x = 1050, y= 600)

        Window2button1 = tk.Button(self, text = 'Intructions', command = open_window, cursor="hand2")
        Window2button1.grid(row= 2, column=3)

        def window_prot2():
            window_prot2 = tk.Toplevel()
            window_prot2.title('Protein')
            window_prot2.geometry('1200x300')
            icon_path = 'Icon Art.png'
            icon = ImageTk.PhotoImage(Image.open(icon_path))
            window_prot2.iconphoto(False, icon)


            reg = LinearRegression()

            x_label = tk.Label(window_prot2, text="Protein_Concentration (mg/mL)(X)")
            x_label.grid(row=0, column=0, sticky="W")

            x_entries = [tk.Entry(window_prot2) for i in range(7)]
            for i, entry in enumerate(x_entries):
                entry.grid(row=0, column=i + 1)

            y_label = tk.Label(window_prot2, text="Absorbance (A.U) (Y)")
            y_label.grid(row=1, column=0, sticky="W")

            y_entries = [tk.Entry(window_prot2) for i in range(7)]
            for i, entry in enumerate(y_entries):
                entry.grid(row=1, column=i + 1)

            slo_entry = tk.Entry(window_prot2)
            slo_entry.grid(row=3, column=2, sticky="W")

            int_entry = tk.Entry(window_prot2)
            int_entry.grid(row=4, column=2, sticky="W")

            def calculate():
                x_values = [entry.get() for entry in x_entries]
                y_values = [entry.get() for entry in y_entries]

                x = []
                y = []
                for i in range(len(x_values)):
                    if x_values[i] and y_values[i]:
                        x.append(float(x_values[i]))
                        y.append(float(y_values[i]))

                if x and y:
                    x = np.array(x).reshape(-1, 1)
                    y = np.array(y)

                    reg.fit(x, y)

                    slope = reg.coef_[0]
                    intercept = reg.intercept_
                    r_squared = reg.score(x, y)

                    slope_value_label.config(text=str(slope))
                    intercept_value_label.config(text=str(intercept))

                    slo_entry.delete(0, 'end')
                    slo_entry.insert(0, str(slope))

                    int_entry.delete(0, 'end')
                    int_entry.insert(0, str(intercept))

                    plt.scatter(x, y)
                    plt.plot(x, reg.predict(x), color='Red')
                    plt.xlabel("Protein Concentration (mg/mL)")
                    plt.ylabel("Absorbance")
                    plt.text(0.05, 0.95, 'R² = {:.3f}'.format(r_squared),
                             transform=plt.gca().transAxes, fontsize=12,
                             verticalalignment='top')
                    plt.show()

                else:
                    messagebox.showerror("Error", "Did you forget the numbers? Decimals must be separated by dot! Do not use comma.")

            calculate_button = tk.Button(window_prot2, text="Calculate", command=calculate)
            calculate_button.grid(row=2, column=1, pady=10)

            slope_label = tk.Label(window_prot2, text="Slope:")
            slope_label.grid(row=3, column=0, sticky="W")

            slope_value_label = tk.Label(window_prot2)
            slope_value_label.grid(row=3, column=1)

            intercept_label = tk.Label(window_prot2, text="Intercept:")
            intercept_label.grid(row=4, column=0, sticky="W")

            intercept_value_label = tk.Label(window_prot2)
            intercept_value_label.grid(row=4, column=1)

            slope_label = tk.Label(window_prot2, text="Sample Absorbance")
            slope_label.grid(row=3, column=4, sticky="W")

            sample_entry = tk.Entry(window_prot2)
            sample_entry.grid(row=4, column=4, sticky="W")

            def sample():
                first = float(sample_entry.get()) - float(int_entry.get())
                second = float(slo_entry.get())
                output = first / second
                print(output)
                output = round(output, 4)
                output_label.config(text=str(output))
                clipboard.copy(output)


            output_label = tk.Label(window_prot2, text='', font=('calibre', 10))
            output_label.grid(row=8, column=5, sticky="W")
            Result_label = tk.Label(window_prot2, text='mg Protein/mL', font=('calibre', 15), bg=("dark gray"))
            Result_label.grid(row=8, column=4, sticky="W")

            Sample_Button = tk.Button(window_prot2, text='Calculate Protein', command=sample)
            Sample_Button.grid(row="4", column="5")

            button_voltar = tk.Button(window_prot2, text='Close', font=('calibre', 20), command=window_prot2.destroy, cursor="hand2")
            button_voltar.place(x=1050, y=600)

        Window_prot2_button1 = tk.Button(self, text='Protein', command=window_prot2, cursor="hand2")
        Window_prot2_button1.grid(row=3, column=3)

class SodWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("SOD Activity Calculator")
        self.geometry("2000x820")
        self.configure(background='#EEFBDC')
        icon_path = 'Icon Art.png'
        icon = ImageTk.PhotoImage(Image.open(icon_path))
        self.iconphoto(False, icon)

        frame1 = tk.LabelFrame(self, text='')
        frame1.pack(padx=1, pady=1)
        frame1.place(relx=0.5, rely=0.3, anchor='c')

        AbsBlank = tk.StringVar()
        AbsSample = tk.StringVar()
        Reaction_volume = tk.StringVar()
        Sample_volume = tk.StringVar()
        Dilution_factor =tk.StringVar()
        mgprot = tk.StringVar()

        def answer():
            RAbs = (float(AbsBlank.get()) - float(AbsSample.get())) / float(AbsBlank.get())
            Unit_Calc = RAbs / 0.5
            React_volume = float(Reaction_volume.get()) / float(Sample_volume.get())
            U =  Unit_Calc * React_volume
            U_mgprot = U / float(mgprot.get())
            output = U_mgprot * float(Dilution_factor.get())
            output = round(output, 4)
            Output_label.config(text=str(output))
            clipboard.copy(output)

        AbsBlankentry = tk.Entry(frame1, width=12, textvariable=AbsBlank)
        AbsSampleentry = tk.Entry(frame1, width=12, textvariable=AbsSample)
        Reaction_volumeentry = tk.Entry(frame1, width=12, textvariable=Reaction_volume)
        Sample_volumeentry = tk.Entry(frame1, width=12, textvariable=Sample_volume)
        mgprotentry = tk.Entry(frame1, width=12, textvariable=mgprot)
        Dilution_factorentry = tk.Entry(frame1, width=12, textvariable=Dilution_factor)

        AbsBlank_label = tk.Label(frame1, text='Absorbance Blank', font=('calibre', 20))
        AbsSample_label = tk.Label(frame1, text='Absorbance Sample', font=('calibre', 20))
        Reaction_volume_label = tk.Label(frame1, text=' Reaction Volume', font=('calibre', 20))
        Sample_volume_label = tk.Label(frame1, text='   Sample Volume', font=('calibre', 20))
        mgprotentry_label = tk.Label(frame1, text='     mg Protein/mL', font=('calibre', 20))
        Dilution_factor_label= tk.Label(frame1, text='     Dilution Factor', font=('calibre', 20))
        Output_label = tk.Label(frame1, text='', font=('calibre', 10))
        Result_label = tk.Label(frame1, text='  Activity: U/mg Protein', font=('calibre', 15), bg=("dark gray"))

        AbsBlank_label.grid(row=2, column=1)
        AbsBlankentry.grid(row=2, column=2)
        AbsSample_label.grid(row=3, column=1)
        AbsSampleentry.grid(row=3, column=2)
        Reaction_volume_label.grid(row=4, column=1)
        Reaction_volumeentry.grid(row=4, column=2)
        Sample_volume_label.grid(row=5, column=1)
        Sample_volumeentry.grid(row=5, column=2)
        mgprotentry_label.grid(row=6, column=1)
        mgprotentry.grid(row=6, column=2)
        Dilution_factor_label.grid(row=7, column=1)
        Dilution_factorentry.grid(row=7, column=2)
        Result_label.grid(row=8, column=1)
        Output_label.grid(row=8, column=2)

        calculate_Button = tk.Button(frame1, text='Calculate', command=answer, cursor="hand2")
        calculate_Button.grid(row="9", column="4")

        self.bind('<Return>', lambda event: calculate_Button.invoke())

        frame2 = tk.LabelFrame(self, text='')
        frame2.pack(padx=5, pady=5)
        frame2.place(relx=0.5, rely=0.7, anchor='c')

        list_types = ["Control", "Sample", "Treated", ""]
        Combobox_select_type1 = tk.StringVar()
        Combobox_select_type1 = ttk.Combobox(frame2, values=list_types, font=("Arial", "8", "bold"))
        Combobox_select_type1.grid(row=1, column=1, columnspan=3)
        Combobox_select_type2 = tk.StringVar()
        Combobox_select_type2 = ttk.Combobox(frame2, values=list_types, font=("Arial", "8", "bold"))
        Combobox_select_type2.grid(row=1, column=4, columnspan=3)
        Combobox_select_type3 = tk.StringVar()
        Combobox_select_type3 = ttk.Combobox(frame2, values=list_types, font=("Arial", "8", "bold"))
        Combobox_select_type3.grid(row=1, column=7, columnspan=3)
        Combobox_select_type4 = tk.StringVar()
        Combobox_select_type4 = ttk.Combobox(frame2, values=list_types, font=("Arial", "8", "bold"))
        Combobox_select_type4.grid(row=1, column=10, columnspan=3)
        Combobox_select_type5 = tk.StringVar()
        Combobox_select_type5 = ttk.Combobox(frame2, values=list_types, font=("Arial", "8", "bold"))
        Combobox_select_type5.grid(row=1, column=13, columnspan=3)
        Combobox_select_type6 = tk.StringVar()
        Combobox_select_type6 = ttk.Combobox(frame2, values=list_types, font=("Arial", "8", "bold"))
        Combobox_select_type6.grid(row=1, column=16, columnspan=3)

        frame2_sampleentry1 = tk.StringVar()
        frame2_sampleentry1 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry1.grid(row=2, column=1)
        frame2_sampleentry2 = tk.StringVar()
        frame2_sampleentry2 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry2.grid(row=2, column=4)
        frame2_sampleentry3 = tk.StringVar()
        frame2_sampleentry3 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry3.grid(row=2, column=7)
        frame2_sampleentry4 = tk.StringVar()
        frame2_sampleentry4 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry4.grid(row=2, column=10)
        frame2_sampleentry5 = tk.StringVar()
        frame2_sampleentry5 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry5.grid(row=2, column=13)
        frame2_sampleentry6 = tk.StringVar()
        frame2_sampleentry6 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry6.grid(row=2, column=16)

        frame2_sampleentry7 = tk.StringVar()
        frame2_sampleentry7 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry7.grid(row=3, column=1)
        frame2_sampleentry8 = tk.StringVar()
        frame2_sampleentry8 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry8.grid(row=3, column=4)
        frame2_sampleentry9 = tk.StringVar()
        frame2_sampleentry9 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry9.grid(row=3, column=7)
        frame2_sampleentry10 = tk.StringVar()
        frame2_sampleentry10 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry10.grid(row=3, column=10)
        frame2_sampleentry11 = tk.StringVar()
        frame2_sampleentry11 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry11.grid(row=3, column=13)
        frame2_sampleentry12 = tk.StringVar()
        frame2_sampleentry12 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry12.grid(row=3, column=16)

        frame2_sampleentry13 = tk.StringVar()
        frame2_sampleentry13 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry13.grid(row=4, column=1)
        frame2_sampleentry14 = tk.StringVar()
        frame2_sampleentry14 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry14.grid(row=4, column=4)
        frame2_sampleentry15 = tk.StringVar()
        frame2_sampleentry15 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry15.grid(row=4, column=7)
        frame2_sampleentry16 = tk.StringVar()
        frame2_sampleentry16 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry16.grid(row=4, column=10)
        frame2_sampleentry17 = tk.StringVar()
        frame2_sampleentry17 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry17.grid(row=4, column=13)
        frame2_sampleentry18 = tk.StringVar()
        frame2_sampleentry18 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry18.grid(row=4, column=16)

        frame2_sampleentry19 = tk.StringVar()
        frame2_sampleentry19 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry19.grid(row=5, column=1)
        frame2_sampleentry20 = tk.StringVar()
        frame2_sampleentry20 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry20.grid(row=5, column=4)
        frame2_sampleentry21 = tk.StringVar()
        frame2_sampleentry21 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry21.grid(row=5, column=7)
        frame2_sampleentry22 = tk.StringVar()
        frame2_sampleentry22 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry22.grid(row=5, column=10)
        frame2_sampleentry23 = tk.StringVar()
        frame2_sampleentry23 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry23.grid(row=5, column=13)
        frame2_sampleentry24 = tk.StringVar()
        frame2_sampleentry24 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry24.grid(row=5, column=16)

        frame2_sampleentry25 = tk.StringVar()
        frame2_sampleentry25 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry25.grid(row=6, column=1)
        frame2_sampleentry26 = tk.StringVar()
        frame2_sampleentry26 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry26.grid(row=6, column=4)
        frame2_sampleentry27 = tk.StringVar()
        frame2_sampleentry27 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry27.grid(row=6, column=7)
        frame2_sampleentry28 = tk.StringVar()
        frame2_sampleentry28 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry28.grid(row=6, column=10)
        frame2_sampleentry29 = tk.StringVar()
        frame2_sampleentry29 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry29.grid(row=6, column=13)
        frame2_sampleentry30 = tk.StringVar()
        frame2_sampleentry30 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry30.grid(row=6, column=16)

        frame2_sampleentry31 = tk.StringVar()
        frame2_sampleentry31 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry31.grid(row=7, column=1)
        frame2_sampleentry32 = tk.StringVar()
        frame2_sampleentry32 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry32.grid(row=7, column=4)
        frame2_sampleentry33 = tk.StringVar()
        frame2_sampleentry33 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry33.grid(row=7, column=7)
        frame2_sampleentry34 = tk.StringVar()
        frame2_sampleentry34 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry34.grid(row=7, column=10)
        frame2_sampleentry35 = tk.StringVar()
        frame2_sampleentry35 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry35.grid(row=7, column=13)
        frame2_sampleentry36 = tk.StringVar()
        frame2_sampleentry36 = tk.Entry(frame2, width=20, font=(15))
        frame2_sampleentry36.grid(row=7, column=16)

        frame3 = tk.LabelFrame(self, text='')
        frame3.pack(padx=10, pady=20)
        frame3.place(relx=0.2, rely=0.95, anchor='sw')

        def open_excel():
            file_path = filedialog.askopenfilename(initialdir="/", title="Select excel file",
                                                   filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
            if file_path.endswith(".xlsx"):
                engine = 'openpyxl'
            elif file_path.endswith(".xls"):
                engine = 'xlrd'
            else:
                raise ValueError("Invalid file extension")

            data = pd.read_excel(file_path, engine=engine)

            # Plot xy graphic using the first and second columns (A and B)
            plt.plot(data.iloc[:, 0], data.iloc[:, 1])
            plt.xlabel('Time (Secs)')
            plt.ylabel('Absorbance (A.U.)')
            plt.ion()
            plt.show()

        open_excel_button = tk.Button(frame3, text='Graph XY Excel', font=("Arial", "12"),
                                      width=16, bg="#99FA13", fg="#000000", activebackground="#286F63",
                                      activeforeground="#D0FEF7", command=open_excel, cursor="hand2")
        open_excel_button.grid(row=1, column=1)

        def plot():
            warnings.filterwarnings("ignore", category=RuntimeWarning)
            group1 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry1, frame2_sampleentry7, frame2_sampleentry13, frame2_sampleentry19,
                       frame2_sampleentry25, frame2_sampleentry31]]

            group2 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry2, frame2_sampleentry8, frame2_sampleentry14, frame2_sampleentry20,
                       frame2_sampleentry26, frame2_sampleentry32]]

            group3 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry3, frame2_sampleentry9, frame2_sampleentry15, frame2_sampleentry21,
                       frame2_sampleentry27, frame2_sampleentry33]]

            group4 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry4, frame2_sampleentry10, frame2_sampleentry16, frame2_sampleentry22,
                       frame2_sampleentry28, frame2_sampleentry34]]

            group5 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry5, frame2_sampleentry11, frame2_sampleentry17, frame2_sampleentry23,
                       frame2_sampleentry29, frame2_sampleentry35]]

            group6 = [float(entry.get()) if entry.get() else None for entry in
                      [frame2_sampleentry6, frame2_sampleentry12, frame2_sampleentry18, frame2_sampleentry24,
                       frame2_sampleentry30, frame2_sampleentry36]]

            def stddev(data):
                n = len(data)
                if n == 0:
                    return np.nan  # Return not a number (NaN) for standard deviation if the group is empty
                mean_value = sum(data) / n
                variance = sum((x - mean_value) ** 2 for x in data) / (n - 1)
                return np.sqrt(variance)

            group1_filtered = [x for x in group1 if x is not None and not np.isnan(x)]
            group1_mean = np.mean(group1_filtered)
            group1_sd = stddev(group1_filtered)

            group2_filtered = [x for x in group2 if x is not None and not np.isnan(x)]
            group2_mean = np.mean(group2_filtered)
            group2_sd = stddev(group2_filtered)

            group3_filtered = [x for x in group3 if x is not None and not np.isnan(x)]
            group3_mean = np.mean(group3_filtered)
            group3_sd = stddev(group3_filtered)

            group4_filtered = [x for x in group4 if x is not None and not np.isnan(x)]
            group4_mean = np.mean(group4_filtered)
            group4_sd = stddev(group4_filtered)

            group5_filtered = [x for x in group5 if x is not None and not np.isnan(x)]
            group5_mean = np.mean(group5_filtered)
            group5_sd = stddev(group5_filtered)

            group6_filtered = [x for x in group6 if x is not None and not np.isnan(x)]
            group6_mean = np.mean(group6_filtered)
            group6_sd = stddev(group6_filtered)

            fig, ax = plt.subplots()

            bar1 = ax.bar(1, group1_mean, yerr=group1_sd, label=Combobox_select_type1.get(), ecolor='red', capsize=5)
            bar2 = ax.bar(2, group2_mean, yerr=group2_sd, label=Combobox_select_type2.get(), ecolor='red', capsize=5)
            bar3 = ax.bar(3, group3_mean, yerr=group3_sd, label=Combobox_select_type3.get(), ecolor='red', capsize=5)
            bar4 = ax.bar(4, group4_mean, yerr=group4_sd, label=Combobox_select_type4.get(), ecolor='red', capsize=5)
            bar5 = ax.bar(5, group5_mean, yerr=group5_sd, label=Combobox_select_type5.get(), ecolor='red', capsize=5)
            bar6 = ax.bar(6, group6_mean, yerr=group6_sd, label=Combobox_select_type6.get(), ecolor='red', capsize=5)



            ax.set_ylabel('Catalase Activity (U/mg protein)')
            ax.set_xticks([1, 2, 3, 4, 5, 6])
            ax.set_xticklabels(
                [(Combobox_select_type1.get()), (Combobox_select_type2.get()), (Combobox_select_type3.get()),
                 (Combobox_select_type4.get()), (Combobox_select_type5.get()), (Combobox_select_type6.get())])

            plt.show()

        plot_button = tk.Button(frame3, text='Plot', font=("Arial", "12"),
                                width=16, bg="#99FA13", fg="#000000", activebackground="#286F63",
                                activeforeground="#D0FEF7", command=plot, cursor="hand2")
        plot_button.grid(row=1, column=2)
        warnings.resetwarnings()


        def clear():
            clear_combobox = [Combobox_select_type1, Combobox_select_type2, Combobox_select_type3,
                              Combobox_select_type4, Combobox_select_type5, Combobox_select_type6,
                              frame2_sampleentry1, frame2_sampleentry2, frame2_sampleentry3, frame2_sampleentry4,
                              frame2_sampleentry5, frame2_sampleentry6, frame2_sampleentry7, frame2_sampleentry8,
                              frame2_sampleentry9, frame2_sampleentry10, frame2_sampleentry11, frame2_sampleentry12,
                              frame2_sampleentry13, frame2_sampleentry14, frame2_sampleentry15, frame2_sampleentry16,
                              frame2_sampleentry17, frame2_sampleentry18, frame2_sampleentry19, frame2_sampleentry20,
                              frame2_sampleentry21, frame2_sampleentry22, frame2_sampleentry23, frame2_sampleentry24,
                              frame2_sampleentry25, frame2_sampleentry26, frame2_sampleentry27, frame2_sampleentry28,
                              frame2_sampleentry29, frame2_sampleentry30, frame2_sampleentry31, frame2_sampleentry32,
                              frame2_sampleentry33, frame2_sampleentry34, frame2_sampleentry35, frame2_sampleentry36]

            result = messagebox.askyesno("Save Changes",
                                         "ARE YOU SURE YOU WANT TO CLEAR? That will erase all unsaved data! ")
            if result == True:
                for widget in clear_combobox:
                    if isinstance(widget, tk.Entry):
                        widget.delete(0, 300)
                    elif isinstance(widget, tk.ttk.Combobox):
                        widget.set("")

        clear_button = tk.Button(frame3, text='Clear', font=("Arial", "12"),
                                 width=16, bg="#99FA13", fg="#000000", activebackground="#286F63",
                                 activeforeground="#D0FEF7", command=clear)
        clear_button.grid(row=1, column=3)

        def save_member():
            widgets = [Combobox_select_type1, Combobox_select_type2, Combobox_select_type3, Combobox_select_type4,
                       Combobox_select_type5, Combobox_select_type6,
                       frame2_sampleentry1, frame2_sampleentry2, frame2_sampleentry3, frame2_sampleentry4,
                       frame2_sampleentry5, frame2_sampleentry6, frame2_sampleentry7, frame2_sampleentry8,
                       frame2_sampleentry9, frame2_sampleentry10, frame2_sampleentry11, frame2_sampleentry12,
                       frame2_sampleentry13, frame2_sampleentry14, frame2_sampleentry15, frame2_sampleentry16,
                       frame2_sampleentry17, frame2_sampleentry18, frame2_sampleentry19, frame2_sampleentry20,
                       frame2_sampleentry21, frame2_sampleentry22, frame2_sampleentry23, frame2_sampleentry24,
                       frame2_sampleentry25, frame2_sampleentry26, frame2_sampleentry27, frame2_sampleentry28,
                       frame2_sampleentry29, frame2_sampleentry30, frame2_sampleentry31, frame2_sampleentry32,
                       frame2_sampleentry33, frame2_sampleentry34, frame2_sampleentry35, frame2_sampleentry36]
            file_name = "SODactivity_" + datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + ".xlsx"
            file = openpyxl.Workbook()
            sheet = file.active

            rows = 7
            cols = 6
            for row in range(rows):
                for col in range(cols):
                    i = row * cols + col
                    sheet.cell(column=col + 2, row=row + 3, value=widgets[i].get())

            file.save(file_name)
            file.close()

        save_button = tk.Button(frame3, text='Save', font=("Arial", "12"),
                                width=15, bg="#99FA13", fg="#000000", activebackground="#286F63",
                                activeforeground="#D0FEF7", command=save_member, cursor="hand2")
        save_button.grid(row= 1, column = 4)

        def openFile():
            the_file = filedialog.askopenfilename(title="Select the updated file", filetypes=[("All files", "*.*")])
            if the_file:
                print(f"Selected file: {the_file}")
                try:
                    os.startfile(os.path.abspath(the_file))
                except Exception as e:
                    print(f"Error opening file: {e}")
            else:
                self.state("iconic")
        open_button = tk.Button(frame3, text="Open File", font=("Arial", "12"),
                                width=18, bg="#99FA13", fg="#000000", activebackground="#286F63",
                                activeforeground="#D0FEF7", command=openFile, cursor="hand2")
        open_button.grid(row= 1, column = 5)


        def open_window():
            window2 = tk.Toplevel()
            window2.title('Instructions')
            window2.geometry('6000x3000')
            icon_path = 'Icon Art.png'
            icon = ImageTk.PhotoImage(Image.open(icon_path))
            window2.iconphoto(False, icon)

            label_name = tk.Label(window2,
                                  text='Absorbance Blank → absorbance of the blank tube - A1 \n'
                                       'Absorbance Sample → absorbance of the sample tube - A2 \n'
                                       'Reaction Volume → total volume in the cuvette \n'
                                       'Samples Volume → volume of samples in the cuvette \n'
                                       'mg protein/mL → protein concentration of the original sample\n'
                                       '\n'
                                       'Formula: SOD U/mL = (((A1-A2)/A1)/ 0.5) x (Reaction volume (mL) /Sample volume (mL)) x DF\n'
		                               '         U/mg protein =  SOD U/mL / mg protein/mL \n'                                   
                                       '\n'
                                       'The result is automatically copied to clipboard so the user may paste in the table provided or in another software.\n'
                                       '\n'
                                       'Users are expected to enter the data in the table and choose or enter the names or codes of the samples.\n'
                                       'Each cell in the table can accept numbers, letters, or symbols. The plot function will not work with letters or symbols. \n'
                                       '\n'
                                       'When pressed, the SAVE BUTTON will save a file in the same folder as the Redoxyme.exe file. \n'
                                       'The file will be saved in Excel format (named "SODactivity year-month-day hour-minute-second.xlsx"). The user should modify the name and save it in a different folder.\n'
                                       'If the SAVE BUTTON is pressed again, it will SAVE IT AGAIN with the current information of Y-M-D H-min-sec.\n'
                                       '\n'
                                       'The user can open the above-saved file or any other file on their computer by pressing the OPEN FILE button. \n'
                                       '\n'
                                       'The CLEAR button, when pressed, will erase all data. Please, check for proper data in the saved file. \n'
                                       '\nContact email: heberty.facundo@gmail.com',
                                        bd=1, justify='left', font=('calibre', 16))

            label_name.grid(row=4, column=0, sticky='w')
            button_voltar = tk.Button(window2, text='Close', font=('calibre', 20), command=window2.destroy, cursor="hand2")
            button_voltar.place(x=1050, y=600)

        Window2button1 = tk.Button(self, text='Intructions', command=open_window, cursor="hand2")
        Window2button1.grid(row=2, column=3)

        def window_prot3():
            window_prot3 = tk.Toplevel()
            window_prot3.title('Protein')
            window_prot3.geometry('1200x300')
            icon_path = 'Icon Art.png'
            icon = ImageTk.PhotoImage(Image.open(icon_path))
            window_prot3.iconphoto(False, icon)

            reg = LinearRegression()

            x_label = tk.Label(window_prot3, text="Protein_Concentration (mg/mL)(X)")
            x_label.grid(row=0, column=0, sticky="W")

            x_entries = [tk.Entry(window_prot3) for i in range(7)]
            for i, entry in enumerate(x_entries):
                entry.grid(row=0, column=i + 1)

            y_label = tk.Label(window_prot3, text="Absorbance (A.U) (Y)")
            y_label.grid(row=1, column=0, sticky="W")

            y_entries = [tk.Entry(window_prot3) for i in range(7)]
            for i, entry in enumerate(y_entries):
                entry.grid(row=1, column=i + 1)

            slo_entry = tk.Entry(window_prot3)
            slo_entry.grid(row=3, column=2, sticky="W")

            int_entry = tk.Entry(window_prot3)
            int_entry.grid(row=4, column=2, sticky="W")

            def calculate():
                x_values = [entry.get() for entry in x_entries]
                y_values = [entry.get() for entry in y_entries]

                x = []
                y = []
                for i in range(len(x_values)):
                    if x_values[i] and y_values[i]:
                        x.append(float(x_values[i]))
                        y.append(float(y_values[i]))

                if x and y:
                    x = np.array(x).reshape(-1, 1)
                    y = np.array(y)

                    reg.fit(x, y)

                    slope = reg.coef_[0]
                    intercept = reg.intercept_
                    r_squared = reg.score(x, y)

                    slope_value_label.config(text=str(slope))
                    intercept_value_label.config(text=str(intercept))

                    slo_entry.delete(0, 'end')
                    slo_entry.insert(0, str(slope))

                    int_entry.delete(0, 'end')
                    int_entry.insert(0, str(intercept))

                    plt.scatter(x, y)
                    plt.plot(x, reg.predict(x), color='Red')
                    plt.xlabel("Protein Concentration (mg/mL)")
                    plt.ylabel("Absorbance")
                    plt.text(0.05, 0.95, 'R² = {:.3f}'.format(r_squared),
                             transform=plt.gca().transAxes, fontsize=12,
                             verticalalignment='top')
                    plt.show()

                else:
                    messagebox.showerror("Error", "Did you forget the numbers? Decimals must be separated by dot! Do not use comma.")

            calculate_button = tk.Button(window_prot3, text="Calculate", command=calculate)
            calculate_button.grid(row=2, column=1, pady=10)

            slope_label = tk.Label(window_prot3, text="Slope:")
            slope_label.grid(row=3, column=0, sticky="W")

            slope_value_label = tk.Label(window_prot3)
            slope_value_label.grid(row=3, column=1)

            intercept_label = tk.Label(window_prot3, text="Intercept:")
            intercept_label.grid(row=4, column=0, sticky="W")

            intercept_value_label = tk.Label(window_prot3)
            intercept_value_label.grid(row=4, column=1)

            slope_label = tk.Label(window_prot3, text="Sample Absorbance")
            slope_label.grid(row=3, column=4, sticky="W")

            sample_entry = tk.Entry(window_prot3)
            sample_entry.grid(row=4, column=4, sticky="W")

            def sample():
                first = float(sample_entry.get()) - float(int_entry.get())
                second = float(slo_entry.get())
                output = first / second
                print(output)
                output = round(output, 4)
                output_label.config(text=str(output))

                clipboard.copy(output)

            output_label = tk.Label(window_prot3, text='', font=('calibre', 10))
            output_label.grid(row=8, column=5, sticky="W")
            Result_label = tk.Label(window_prot3, text='mg Protein/mL', font=('calibre', 15), bg=("dark gray"))
            Result_label.grid(row=8, column=4, sticky="W")

            Sample_Button = tk.Button(window_prot3, text='Calculate Protein', command=sample)
            Sample_Button.grid(row="4", column="5")

            button_voltar = tk.Button(window_prot3, text='Close', font=('calibre', 20), command=window_prot3.destroy, cursor="hand2")
            button_voltar.place(x=1050, y=600)

        Window_prot3_button1 = tk.Button(self, text='Protein', command=window_prot3, cursor="hand2")
        Window_prot3_button1.grid(row=3, column=3)


if __name__ == "__main__":
    root = MainWindow()
    root.mainloop()